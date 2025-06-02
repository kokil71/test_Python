# python으로 엑셀에서 범위를 선택하고 캐드에 그것을 표로 그리는 코드

from pyautocad import Autocad, APoint
from openpyxl import load_workbook

def read_excel_range(file_path, sheet_name, range_str):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    data = []
    for row in ws[range_str]:
        data.append([cell.value if cell.value is not None else "" for cell in row])
    return data

def draw_table_in_autocad(data, start_point=(0, 0), row_height=5, col_width=20):
    acad = Autocad(create_if_not_exists=True)
    x0, y0 = start_point

    for i, row in enumerate(data):
        for j, value in enumerate(row):
            x = x0 + j * col_width
            y = y0 - i * row_height
            # Draw rectangle cell (optional)
            p1 = APoint(x, y)
            p2 = APoint(x + col_width, y - row_height)
            acad.model.AddLine(p1, APoint(p2.x, p1.y))  # top
            acad.model.AddLine(p1, APoint(p1.x, p2.y))  # left
            acad.model.AddLine(p2, APoint(p2.x, p1.y))  # right
            acad.model.AddLine(p2, APoint(p1.x, p2.y))  # bottom

            # Add text
            text_point = APoint(x + col_width / 2, y - row_height / 2)
            acad.model.AddText(str(value), text_point, 2.5)

# 사용 예시
excel_file = r"d:\work\_test_data.xlsx"
sheet = "AutoCAD_Export"
cell_range = "A1:E4"

data = read_excel_range(excel_file, sheet, cell_range)
draw_table_in_autocad(data, start_point=(0, 0))
