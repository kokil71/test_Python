# python으로 엑셀에서 범위를 선택하고 캐드에 그것을 표로 그리는 코드

from pyautocad import Autocad, APoint
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

class ExcelTable:
    def __init__(self, file_path, sheet_name, cell_range):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.cell_range = cell_range
        self.data, self.alignments = self._read_excel_data()

    def _read_excel_data(self):
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb[self.sheet_name]

        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)
        data = []
        alignments = []

        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            data_row = []
            align_row = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                data_row.append(str(value))

                align = 'left'  # 기본 정렬
                if cell.alignment and cell.alignment.horizontal:
                    align = cell.alignment.horizontal
                align_row.append(align)
            data.append(data_row)
            alignments.append(align_row)

        return data, alignments

    def draw_in_cad(self, start_point=(0, 0), row_height=5, col_width=20,
                    text_height=2.5, text_style_name="HangulTTF"):
        acad = Autocad(create_if_not_exists=True)
        doc = acad.doc

        # 텍스트 스타일 설정 (한글 폰트용)
        styles = doc.TextStyles
        try:
            styles.Item(text_style_name)
        except:
            style = styles.Add(text_style_name)
            style.FontFile = "malgun.ttf"  # 맑은 고딕

        x0, y0 = start_point

        for i, (row, align_row) in enumerate(zip(self.data, self.alignments)):
            for j, (text, align) in enumerate(zip(row, align_row)):
                x = x0 + j * col_width
                y = y0 - i * row_height

                # 셀 테두리
                p1 = APoint(x, y)
                p2 = APoint(x + col_width, y - row_height)
                acad.model.AddLine(p1, APoint(p2.x, p1.y))  # top
                acad.model.AddLine(p1, APoint(p1.x, p2.y))  # left
                acad.model.AddLine(p2, APoint(p2.x, p1.y))  # right
                acad.model.AddLine(p2, APoint(p1.x, p2.y))  # bottom

                # 정렬 좌표 설정
                if align == 'center':
                    tx = x + col_width / 2 - len(text) * text_height * 0.25
                elif align == 'right':
                    tx = x + col_width - len(text) * text_height * 0.5 - 1
                else:  # left or None
                    tx = x + 1
                ty = y - row_height + 1

                txt = acad.model.AddText(text, APoint(tx, ty), text_height)
                txt.StyleName = text_style_name

# 사용 예시
if __name__ == "__main__":
    excel_file = r"d:\work\_test_data.xlsx"
    sheet = "AutoCAD_Export"
    cell_range = "A1:E4"
    #table = ExcelTable("sample.xlsx", "Sheet1", "A1:D5")
    table = ExcelTable(excel_file, sheet, cell_range)
    table.draw_in_cad(start_point=(0, 0))
