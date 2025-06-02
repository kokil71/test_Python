# 엑셀에서 Data 읽어와서 캐드 표 만드는 코드
# 캐드 표를 엑셀로 만들기

import win32com.client
import pythoncom
import openpyxl
import os
import sys
import time
from openpyxl import Workbook

# 공통 유틸
def APoint(x, y, z=0):
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))

def aDouble(*xyz):
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, xyz)

def left(s: str, n: int) -> str:
    return s[:n]

def create_acad_color(acad_app, r, g, b):
    version = left(acad_app.ActiveDocument.Application.Version, 2)
    color = acad_app.ActiveDocument.Application.GetInterfaceObject(f"AutoCAD.AcCmColor.{version}")
    color.SetRGB(r, g, b)
    return color

# 엑셀 읽기
def read_excel_data(file_path: str, sheet_name: str = None) -> list:
    if not os.path.exists(file_path):
        print(f"❌ 엑셀 파일 없음: {file_path}")
        sys.exit(1)
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active
        data = [list(row) for row in sheet.iter_rows(values_only=True) if any(cell is not None for cell in row)]
        return data
    except Exception as e:
        print(f"🚫 엑셀 읽기 오류: {e}")
        sys.exit(1)

# 엑셀 저장
def write_excel_data1(data: list, file_path: str):
    try:
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(file_path)
        print(f"📁 엑셀 저장 완료: {file_path}")
    except Exception as e:
        print(f"🚫 엑셀 저장 오류: {e}")

# AutoCAD 연결
def connect_autocad():
    try:
        acad = win32com.client.GetActiveObject("AutoCAD.Application")
    except:
        acad = win32com.client.Dispatch("AutoCAD.Application")
        acad.Visible = True
        time.sleep(3)
    return acad

# 캐드에 표 생성
def draw_table(acad_app, data: list, insert_point=(0, 0, 0), row_height=5, column_width=30):
    if not data or not data[0]:
        print("❌ 유효한 데이터 없음.")
        return

    doc = acad_app.ActiveDocument
    ms = doc.ModelSpace

    insert_point = APoint(*insert_point)
    rows = len(data)
    cols = max(len(row) for row in data)

    try:
        #table = ms.AddTable(insert_point, rows+2, cols, row_height, column_width)      # 타이틀 / 헤더 삭제시
        table = ms.AddTable(insert_point, rows+1, cols, row_height, column_width)       # 타이틀 만 삭제시
        table.StyleName = "Standard"
        table.Direction = aDouble(1, 0, 0)

        # 타이틀/헤더 제거 시도
        try:
            #table.DeleteRows(0, 2)                     # 타이틀 / 헤더 둘다 삭제시
            table.DeleteRows(0, 1)                      # 타이틀 만 삭제시
        except:
            pass

        for r in range(rows):
            for c in range(cols):
                value = data[r][c] if c < len(data[r]) else ""
                table.SetText(r, c, str(value))
                table.SetCellAlignment(r, c, 5)
                table.SetCellTextHeight(r, c, 2)

        print("✅ AutoCAD 테이블 생성 완료")

    except Exception as e:
        print(f"🚫 테이블 생성 오류: {e}")


def select_entity1(acad):
    doc = acad.ActiveDocument

    # SelectionSet 초기화 또는 새로 생성
    try:
        ss = doc.SelectionSets.Item("SS1")
        ss.Clear()
    except:
        ss = doc.SelectionSets.Add("SS1")

    print("📌 AutoCAD에서 표(테이블)를 선택하세요 (하나 선택 후 Enter):")
    ss.SelectOnScreen()

    if ss.Count == 0:
        print("⚠️ 선택된 객체가 없습니다.")
        return None

    entity = ss.Item(0)
    # 선택한 객체가 표인지 확인
    if entity.ObjectName != "AcDbTable":
        print(f"⚠️ 선택된 객체가 표가 아닙니다. 객체 유형: {entity.ObjectName}")
        return None

    return entity

# AutoCAD 표 선택 후 Excel 저장
def export_table_to_excel1(acad_app, save_path: str):
    
        # 엔터티와 픽 포인트를 위한 VARIANT 참조 객체
        # GetEntity는 byref 변수를 직접 수정하므로, `VARIANT`로 감싼 배열을 사용하지 말고 빈 리스트로 참조 넘김
        obj_holder = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_DISPATCH, None)
        point_holder = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_ARRAY | pythoncom.VT_R8, (0.0, 0.0, 0.0))

    #try:
        doc = acad_app.ActiveDocument
        #selection = doc.Utility.GetEntity("📌 AutoCAD 표를 선택하세요", "")
        #doc.Utility.GetEntity(picked_obj, pick_point, "📌 AutoCAD 표를 선택하세요: ")
        doc.Utility.GetEntity(obj_holder[0], point_holder, "📌 AutoCAD 표를 선택하세요: ")
        table = obj_holder.value

        table = select_entity(acad_app)

        if table.ObjectName != "AcDbTable":
            print("❌ 선택된 객체가 표가 아닙니다.")
            return
                
        # 엑셀로 저장
        wb = openpyxl.Workbook()
        ws = wb.active

        rows = table.Rows
        cols = table.Columns

        for r in range(rows):
            for c in range(cols):
                try:
                    value = table.GetText(r, c)
                    ws.cell(row=r + 1, column=c + 1).value = value
                except Exception as e:
                    print(f"⚠️ 셀 ({r},{c}) 읽기 실패: {e}")

        wb.save(export_file)
        print(f"✅ 엑셀로 저장 완료: {export_file}")

        """
        rows = table.Rows
        cols = table.Columns
        data = []

        for r in range(rows):
            row_data = []
            for c in range(cols):
                txt = table.GetText(r, c)
                row_data.append(txt)
            data.append(row_data)

        write_excel_data(data, save_path)
        """

    #except Exception as e:
    #    print(f"🚫 표 추출 오류: {e}")

def select_table(acad):
    doc = acad.ActiveDocument

    # SelectionSet 초기화 또는 새로 생성
    try:
        ss = doc.SelectionSets.Item("SS1")
        ss.Clear()
    except:
        ss = doc.SelectionSets.Add("SS1")

    print("📌 AutoCAD에서 표(테이블)를 선택하세요 (하나 선택 후 Enter):")
    ss.SelectOnScreen()

    if ss.Count == 0:
        print("⚠️ 선택된 객체가 없습니다.")
        return None

    entity = ss.Item(0)
    # 선택한 객체가 표인지 확인
    if entity.ObjectName != "AcDbTable":
        print(f"⚠️ 선택된 객체가 표가 아닙니다. 객체 유형: {entity.ObjectName}")
        return None

    return entity

def export_table_to_excel1(table, excel_path):
    # 엑셀 워크북, 시트 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AutoCAD Table Data"

    rows = table.Rows
    cols = table.Columns

    for r in range(rows):
        for c in range(cols):
            text = table.GetText(r, c)
            ws.cell(row=r+1, column=c+1, value=text)

    # 저장
    wb.save(excel_path)
    print(f"✅ 엑셀 파일 저장 완료: {excel_path}")


def export_table_to_excel2(table, excel_path):
    # 엑셀 파일 존재 여부 체크
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        print(f"📁 기존 파일 열기: {excel_path}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AutoCAD Table Data"
        print(f"🆕 새 파일 생성: {excel_path}")

    rows = table.Rows
    cols = table.Columns

    for r in range(rows):
        for c in range(cols):
            text = table.GetText(r, c)
            ws.cell(row=r+1, column=c+1, value=text)

    wb.save(excel_path)
    print(f"✅ 엑셀 파일 저장 완료: {excel_path}")

def export_table_to_excel(table, excel_path):
    folder = os.path.dirname(excel_path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder)

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        print(f"📁 기존 파일 열기: {excel_path}")
    else:
        wb = openpyxl.Workbook()
        print(f"🆕 새 파일 생성: {excel_path}")

    ws = wb.active

    rows = table.Rows
    cols = table.Columns

    for r in range(rows):
        for c in range(cols):
            text = table.GetText(r, c)
            ws.cell(row=r+1, column=c+1, value=text)

    wb.save(excel_path)
    print(f"✅ 엑셀 파일 저장 완료: {excel_path}")
"""
# 메인
if __name__ == "__main__":
    pythoncom.CoInitialize()

    excel_file = r"D:\work\makeTable.xlsx"
    export_file = r"D:\work\exportedTable.xlsx"

    acad = connect_autocad()

    # 1. 엑셀 → AutoCAD 표 생성
    if os.path.exists(excel_file):
        data = read_excel_data(excel_file)
        draw_table(acad, data)
    else:
        print("📂 엑셀 파일이 없어서 AutoCAD 표 생성 생략")

    # 2. AutoCAD 표 선택 → 엑셀 저장
    table = select_table(acad)
    if table:
        #export_file = os.path.join(os.path.expanduser("~"), "Desktop", "exported_table.xlsx")
        export_table_to_excel(table, export_file)        
    #export_table_to_excel(acad, export_file)

    pythoncom.CoUninitialize()
"""