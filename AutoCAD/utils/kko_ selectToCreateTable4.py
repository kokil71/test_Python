# python으로 엑셀에서 범위를 선택하고 캐드에 그것을 표로 그리는 코드

from pyautocad import Autocad, APoint
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

class ExcelToCADTable:
    def __init__(self, excel_file, sheet_name, cell_range):
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.cell_range = cell_range
        self.data, self.alignments = self._read_excel()
        self.col_widths = self._calculate_col_widths()
        self.row_height = 5
        self.base_text_height = 2.5

    def _read_excel(self):
        wb = load_workbook(self.excel_file, data_only=True)
        ws = wb[self.sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)

        data, aligns = [], []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            drow, arow = [], []
            for cell in row:
                text = str(cell.value) if cell.value is not None else ""
                align = cell.alignment.horizontal or 'left'
                drow.append(text)
                arow.append(align)
            data.append(drow)
            aligns.append(arow)
        return data, aligns

    def _calculate_col_widths(self):
        # 열마다 가장 긴 텍스트 길이 기준으로 폭 결정
        col_count = len(self.data[0])
        widths = [0] * col_count
        char_width = 1.2  # 평균 한 글자당 폭 비율

        for row in self.data:
            for i, cell in enumerate(row):
                length = len(cell)
                widths[i] = max(widths[i], length * char_width + 4)  # 여백 포함

        return widths

    def draw_in_cad(self, start_point=(0, 0), font_name="malgun.ttf", style_name="HangulTTF"):
        acad = Autocad(create_if_not_exists=True)
        doc = acad.doc

        # 텍스트 스타일 설정
        try:
            doc.TextStyles.Item(style_name)
        except:
            style = doc.TextStyles.Add(style_name)
            style.FontFile = font_name

        x0, y0 = start_point

        for row_idx, row in enumerate(self.data):
            y = y0 - row_idx * self.row_height
            for col_idx, text in enumerate(row):
                x = x0 + sum(self.col_widths[:col_idx])
                width = self.col_widths[col_idx]

                # 셀 테두리
                p1 = APoint(x, y)
                p2 = APoint(x + width, y - self.row_height)
                acad.model.AddLine(p1, APoint(p2.x, p1.y))  # top
                acad.model.AddLine(p1, APoint(p1.x, p2.y))  # left
                acad.model.AddLine(p2, APoint(p2.x, p1.y))  # right
                acad.model.AddLine(p2, APoint(p1.x, p2.y))  # bottom

                # 텍스트 위치와 크기 계산
                text_len = max(len(text), 1)
                max_chars_fit = int(width / 1.2)
                scale_factor = min(1.0, max_chars_fit / text_len)
                text_height = self.base_text_height * scale_factor

                # 정렬별 X 위치
                align = self.alignments[row_idx][col_idx]
                if align == 'center':
                    tx = x + width / 2 - text_len * text_height * 0.25
                elif align == 'right':
                    tx = x + width - text_len * text_height * 0.5 - 1
                else:  # left
                    tx = x + 1

                ty = y - self.row_height + 1

                text_obj = acad.model.AddText(text, APoint(tx, ty), text_height)
                text_obj.StyleName = style_name

# 🧪 사용 예시
if __name__ == "__main__":
    excel_file = r"d:\work\_test_data.xlsx"
    sheet = "AutoCAD_Export"
    cell_range = "A1:F4"
    #table = ExcelToCADTable("sample.xlsx", "Sheet1", "A1:D10")
    table = ExcelToCADTable(excel_file, sheet, cell_range)    
    table.draw_in_cad(start_point=(0, 0))
