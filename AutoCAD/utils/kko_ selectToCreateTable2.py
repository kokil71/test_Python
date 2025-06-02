# python으로 엑셀에서 범위를 선택하고 캐드에 그것을 표로 그리는 코드

from pyautocad import Autocad, APoint
from openpyxl import load_workbook

def read_excel_range(file_path, sheet_name, range_str):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    data = []
    for row in ws[range_str]:
        data.append([cell.value if cell.value is not None else "" for cell in row])
    return data

def draw_table_in_autocad(data, start_point=(0, 0), row_height=5, col_width=20, text_height=2.5, text_style_name="HanguelText"):
    acad = Autocad(create_if_not_exists=True)
    doc = acad.doc

    # 텍스트 스타일 설정 또는 생성 (한글용 TTF)
    text_styles = doc.TextStyles
    try:
        text_style = text_styles.Item(text_style_name)
    except:
        # 스타일이 없다면 생성
        text_style = text_styles.Add(text_style_name)
        text_style.FontFile = "gulim.ttf"  # 윈도우에 있는 한글 폰트
        text_style.BigFontFile = ""  # 필요시 SHX big font 파일 지정 가능

    x0, y0 = start_point

    for i, row in enumerate(data):
        for j, value in enumerate(row):
            x = x0 + j * col_width
            y = y0 - i * row_height

            # 사각형 테두리
            p1 = APoint(x, y)
            p2 = APoint(x + col_width, y - row_height)
            acad.model.AddLine(p1, APoint(p2.x, p1.y))  # top
            acad.model.AddLine(p1, APoint(p1.x, p2.y))  # left
            acad.model.AddLine(p2, APoint(p2.x, p1.y))  # right
            acad.model.AddLine(p2, APoint(p1.x, p2.y))  # bottom

            # 한글 텍스트
            text_point = APoint(x + 1, y - row_height + 1)  # padding
            text = acad.model.AddText(str(value), text_point, text_height)
            text.StyleName = text_style_name

# 사용 예시
#excel_file = "sample.xlsx"
#sheet = "Sheet1"
#cell_range = "A1:D5"
excel_file = r"d:\work\_test_data.xlsx"
sheet = "AutoCAD_Export"
cell_range = "A1:E4"

data = read_excel_range(excel_file, sheet, cell_range)
draw_table_in_autocad(data)
