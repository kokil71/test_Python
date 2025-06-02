# python으로 엑셀에서 범위를 선택하고 캐드에 그것을 표로 그리는 코드

import comtypes.client
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from pyautocad import Autocad, APoint


def get_display_texts(excel_path, sheet_name, cell_range):
    excel = comtypes.client.CreateObject("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(excel_path)
    ws = wb.Sheets(sheet_name)

    rng = ws.Range(cell_range)
    display_data = []

    for i in range(1, rng.Rows.Count + 1):
        row_data = []
        for j in range(1, rng.Columns.Count + 1):
            cell = rng.Cells(i, j)
            row_data.append(cell.Text)
        display_data.append(row_data)

    wb.Close(SaveChanges=False)
    excel.Quit()
    return display_data


def get_alignment_and_merge_info(excel_path, sheet_name, cell_range):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # 병합 셀 정보
    merged_map = {}
    merge_sizes = {}

    for r in ws.merged_cells.ranges:
        minc, minr, maxc, maxr = r.bounds
        for row in range(minr, maxr + 1):
            for col in range(minc, maxc + 1):
                merged_map[(row, col)] = (minr, minc)
        merge_sizes[(minr, minc)] = (maxr - minr + 1, maxc - minc + 1)

    alignments = []
    for i, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                         min_col=min_col, max_col=max_col), start=min_row):
        arow = []
        for j, cell in enumerate(row, start=min_col):
            if (i, j) in merged_map and (i, j) != merged_map[(i, j)]:
                arow.append(None)
                continue
            align = cell.alignment.horizontal or "left"
            arow.append(align)
        alignments.append(arow)

    return alignments, merge_sizes


class ExcelToAutoCADTable:
    def __init__(self, excel_path, sheet_name, cell_range):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.cell_range = cell_range

        self.data = get_display_texts(excel_path, sheet_name, cell_range)
        self.alignments, self.merged_info = get_alignment_and_merge_info(excel_path, sheet_name, cell_range)

        self.col_widths = self._calculate_col_widths()
        self.row_height = 5
        self.base_text_height = 2.5

    def _calculate_col_widths(self):
        col_count = len(self.data[0])
        widths = [0] * col_count
        char_width = 1.2

        for row in self.data:
            for i, cell in enumerate(row):
                if cell is not None:
                    length = len(cell)
                    widths[i] = max(widths[i], length * char_width + 4)
        return widths

    def draw_in_cad(self, start_point=(0, 0), style_name="HangulTTF", font_file="malgun.ttf"):
        acad = Autocad(create_if_not_exists=True)
        doc = acad.doc

        try:
            doc.TextStyles.Item(style_name)
        except:
            style = doc.TextStyles.Add(style_name)
            style.FontFile = font_file

        x0, y0 = start_point

        for row_idx, row in enumerate(self.data):
            y = y0 - row_idx * self.row_height

            for col_idx, text in enumerate(row):
                if text is None:
                    continue

                x = x0 + sum(self.col_widths[:col_idx])
                width = self.col_widths[col_idx]

                row_abs = row_idx + 1
                col_abs = col_idx + 1
                rowspan, colspan = 1, 1
                if (row_abs, col_abs) in self.merged_info:
                    rowspan, colspan = self.merged_info[(row_abs, col_abs)]

                total_width = sum(self.col_widths[col_idx:col_idx + colspan])
                total_height = self.row_height * rowspan

                # 셀 테두리
                p1 = APoint(x, y)
                p2 = APoint(x + total_width, y - total_height)
                acad.model.AddLine(p1, APoint(p2.x, p1.y))  # 상
                acad.model.AddLine(p1, APoint(p1.x, p2.y))  # 좌
                acad.model.AddLine(APoint(p2.x, p1.y), p2)  # 우
                acad.model.AddLine(APoint(p1.x, p2.y), p2)  # 하

                # 글자 위치 및 크기 조정
                align = self.alignments[row_idx][col_idx]
                max_text_width = total_width - 2
                text_len = max(len(text), 1)
                ideal_text_width = text_len * 0.6
                scale = min(1.0, max_text_width / ideal_text_width)
                text_height = self.base_text_height * scale

                if align == "center":
                    tx = x + total_width / 2 - (text_len * text_height * 0.25)
                elif align == "right":
                    tx = x + total_width - (text_len * text_height * 0.5) - 1
                else:
                    tx = x + 1
                ty = y - total_height + 1

                t = acad.model.AddText(text, APoint(tx, ty), text_height)
                t.StyleName = style_name


# 🧪 사용 예시
if __name__ == "__main__":
    excel_file = r"d:\work\_test_data.xlsx"
    sheet = "AutoCAD_Export"
    cell_range = "A1:F4"
    #table = ExcelToAutoCADTable("sample.xlsx", "Sheet1", "A1:D10")
    table = ExcelToAutoCADTable(excel_file, sheet, cell_range)    
    table.draw_in_cad(start_point=(0, 0))