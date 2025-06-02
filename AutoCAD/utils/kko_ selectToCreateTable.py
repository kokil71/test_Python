# python으로 엑셀에서 범위를 선택하고 캐드에 그것을 표로 그리는 코드

import os
import comtypes.client
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from pyautocad import Autocad, APoint, ACAD
#from comtypes import VARIANT
from comtypes.automation import VARIANT, VT_ARRAY, VT_R8
from ctypes import byref

def get_display_texts(excel_path, sheet_name, cell_range):
    excel = comtypes.client.CreateObject("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(excel_path)
    ws = wb.Sheets(sheet_name)

    rng = ws.Range(cell_range)
    display_data = []

    for i in range(1, rng.Rows.Count + 1):
        row_data = []
        for j in range(1, rng.Columns.Count + 1):
            cell = rng.Cells(i, j)
            row_data.append(cell.Text)
        display_data.append(row_data)

    wb.Close(SaveChanges=False)
    del ws
    del wb
    excel.Quit()
    return display_data


def get_alignment_and_merge_info(excel_path, sheet_name, cell_range):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    merged_map = {}
    merge_sizes = {}

    for r in ws.merged_cells.ranges:
        minc, minr, maxc, maxr = r.bounds
        for row in range(minr, maxr + 1):
            for col in range(minc, maxc + 1):
                merged_map[(row, col)] = (minr, minc)
        merge_sizes[(minr, minc)] = (maxr - minr + 1, maxc - minc + 1)

    alignments = []
    for i, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                         min_col=min_col, max_col=max_col), start=min_row):
        arow = []
        for j, cell in enumerate(row, start=min_col):
            if (i, j) in merged_map and (i, j) != merged_map[(i, j)]:
                arow.append(None)
                continue
            align = cell.alignment.horizontal or "left"
            arow.append(align)
        alignments.append(arow)

    return alignments, merge_sizes


"""
# 폰트 스타일 목록 가져오기
text_styles = acad.doc.TextStyles

# 폰트 스타일이 존재하는지 확인
if style_name in [text_styles.Item(i).Name for i in range(text_styles.Count)]:
    print(f"폰트 스타일 '{style_name}'이 이미 존재합니다. 그대로 사용합니다.")
else:
    print(f"폰트 스타일 '{style_name}'이 존재하지 않습니다. 새롭게 생성합니다.")
    text_styles.Add(style_name)
    
    # 생성된 폰트 스타일 속성 설정
    new_style = text_styles.Item(style_name)
    new_style.FontFile = "Arial.ttf"  # 사용할 폰트 파일
    new_style.Height = 2.5  # 기본 높이
    new_style.WidthFactor = 1.0  # 너비 비율
    new_style.ObliqueAngle = 0  # 기울기 없음

print("폰트 스타일 설정 완료!")
"""

class ExcelToAutoCADTable:
    def __init__(self, excel_path, sheet_name, cell_range):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.cell_range = cell_range

        self.data = get_display_texts(excel_path, sheet_name, cell_range)
        self.alignments, self.merged_info = get_alignment_and_merge_info(excel_path, sheet_name, cell_range)

        self.acad = Autocad(create_if_not_exists=True)
        self.doc = self.acad.doc

        font_path = os.path.join(os.environ["WINDIR"], "Fonts")
        #print(f"Windows 폰트 폴더 경로: {font_path}")
        self.style_name = "HangulTTF"
        #self.font_file = "malgun.ttf"
        #self.font_file = r"C:\Windows\Fonts\malgun.ttf"
        #self.font_file = r"C:\Windows\Fonts\h2gtre.ttf"
        #self.font_file = fr"{font_path}\malgun.ttf"
        self.font_file = fr"{font_path}\h2gtre.ttf"
        #self.font_file = "gulim.ttf"  # 윈도우에 있는 한글 폰트
        #self.font_file = "gulim.ttc"  # 윈도우에 있는 한글 폰트
        #self.font_file = "Arial.ttf"   # 폰트 이름
        self.base_text_height = 2.5
        self.cell_padding = 1.5

        self._add_text_style()    

    def _add_text_style(self):
        try:
            style = self.doc.TextStyles.Item(self.style_name)
            if style.FontFile != self.font_file:
                #style.FontFile = f"C:\Windows\Fonts\{self.font_file}"
                style.FontFile = self.font_file
        except Exception:
            style = self.doc.TextStyles.Add(self.style_name)
            #style.FontFile = f"C:\Windows\Fonts\{self.font_file}"
            style.FontFile = self.font_file
            #print(f"폰트 파일: {style.FontFile}")
        finally:
            #print(f"폰트 파일: {self.font_file}")
            print(f"폰트 파일: {style.FontFile}")

    def _get_text_size(self, text, height):
        if not text.strip():
            return 0, 0  # 빈 텍스트는 크기 없음

        temp_text = self.acad.model.AddText(text, APoint(0, 0), height)
        temp_text.StyleName = self.style_name

        min_point = VARIANT(VT_ARRAY | VT_R8, [0.0, 0.0, 0.0])
        max_point = VARIANT(VT_ARRAY | VT_R8, [0.0, 0.0, 0.0])

        try:
            temp_text.GetBoundingBox(byref(min_point), byref(max_point))
        except Exception as e:
            temp_text.Delete()
            print(f"⚠️ BoundingBox 실패 (텍스트: '{text}') - {e}")
            return 0, 0

        min_vals = min_point.value
        max_vals = max_point.value

        width = max_vals[0] - min_vals[0]
        height = max_vals[1] - min_vals[1]

        temp_text.Delete()
        
        print(f"글자/폭/높이: {text}, {width}, {height} ")

        return width, height

    def _calculate_cell_sizes(self):
        row_count = len(self.data)
        col_count = len(self.data[0])

        col_widths = [0] * col_count
        row_heights = [0] * row_count

        cell_sizes = [[(0, 0) for _ in range(col_count)] for _ in range(row_count)]

        for r in range(row_count):
            for c in range(col_count):
                text = self.data[r][c]
                if text is None:
                    continue
                w, h = self._get_text_size(text, self.base_text_height)
                w += self.cell_padding * 2
                h += self.cell_padding * 2
                cell_sizes[r][c] = (w, h)

        for r in range(row_count):
            for c in range(col_count):
                if self.data[r][c] is None:
                    continue
                abs_r = r + 1
                abs_c = c + 1
                if (abs_r, abs_c) in self.merged_info:
                    rowspan, colspan = self.merged_info[(abs_r, abs_c)]
                else:
                    rowspan, colspan = 1, 1

                total_width = sum(cell_sizes[r][c + i][0] for i in range(colspan) if c + i < col_count)
                total_height = sum(cell_sizes[r + i][c][1] for i in range(rowspan) if r + i < row_count)

                avg_width = total_width / colspan
                avg_height = total_height / rowspan

                for i in range(colspan):
                    if c + i < col_count:
                        col_widths[c + i] = max(col_widths[c + i], avg_width)
                for i in range(rowspan):
                    if r + i < row_count:
                        row_heights[r + i] = max(row_heights[r + i], avg_height)

        self.col_widths = col_widths
        self.row_heights = row_heights

    def draw_in_cad(self, start_point=(0, 0)):
        self._calculate_cell_sizes()

        x0, y0 = start_point
        row_count = len(self.data)
        col_count = len(self.data[0])

        for r in range(row_count):
            y_top = y0 - sum(self.row_heights[:r])
            row_height = self.row_heights[r]

            for c in range(col_count):
                if self.data[r][c] is None:
                    continue

                x_left = x0 + sum(self.col_widths[:c])
                col_width = self.col_widths[c]

                abs_r = r + 1
                abs_c = c + 1
                if (abs_r, abs_c) in self.merged_info:
                    rowspan, colspan = self.merged_info[(abs_r, abs_c)]
                else:
                    rowspan, colspan = 1, 1

                total_width = sum(self.col_widths[c:c + colspan])
                total_height = sum(self.row_heights[r:r + rowspan])

                # 사각형 테두리
                p1 = APoint(x_left, y_top)
                p2 = APoint(x_left + total_width, y_top - total_height)
                self.acad.model.AddLine(p1, APoint(p2.x, p1.y))
                self.acad.model.AddLine(p1, APoint(p1.x, p2.y))
                self.acad.model.AddLine(APoint(p2.x, p1.y), p2)
                self.acad.model.AddLine(APoint(p1.x, p2.y), p2)

                align = self.alignments[r][c] or "left"
                text = self.data[r][c]

                text_w, text_h = self._get_text_size(text, self.base_text_height)

                # acAlignmentMiddleLeft = 9 / acAlignmentMiddleCenter = 10 / acAlignmentMiddleRight = 11
                if align == "center":
                    #tx = x_left + (total_width - text_w) / 2
                    tx = x_left + total_width / 2                    
                    lngAlign = ACAD.acAlignmentMiddleCenter #lngAlign = 10
                elif align == "right":
                    #tx = x_left + total_width - text_w - self.cell_padding
                    tx = x_left + total_width - self.cell_padding
                    lngAlign = ACAD.acAlignmentMiddleRight  #lngAlign = 11
                    #print(f"tx = x_left + total_width - text_w - self.cell_padding: {tx:.3f} = {x_left:.3f} + {total_width:.3f} - {text_w:.3f} - {self.cell_padding:.3f}")
                else:
                    tx = x_left + self.cell_padding                    
                    lngAlign = ACAD.acAlignmentMiddleLeft   #lngAlign = 9

                #ty = y_top - total_height + (total_height - text_h) / 2 + text_h
                ty = y_top - total_height / 2
                
                #print(f"좌표_이전: {tx} ,  {ty}")

                t = self.acad.model.AddText(text, APoint(tx, ty), self.base_text_height)                
                t.StyleName = self.style_name                
                t.Alignment = lngAlign
                t.TextAlignmentPoint = APoint(tx, ty)  # 원하는 위치로 재설정                
                #print(f"TextAlignmentPoint 좌표: {t.TextAlignmentPoint}")
                #print(f"좌표_이후: {tx} ,  {ty}")                                
                t.Update()  # 변경 적용
                #self.acad.app.Update()

                #print(f"수평 정렬: {align} ,  {lngAlign}")
        
        
        # 화면 갱신
        #self.acad.app.Update()
        self.acad.app.ZoomExtents()
        print("✅ AutoCAD 표 작성 완료")


# ===== 예제 사용 =====
if __name__ == "__main__":
    excel_file = r"d:\work\__test_data.xlsx"
    sheet = "AutoCAD_Export"
    cell_range = "A1:F4"
    table = ExcelToAutoCADTable(excel_file, sheet, cell_range)    
    table.draw_in_cad(start_point=(0, 0))