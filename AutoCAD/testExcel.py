import openpyxl
import comtypes.client
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from pyautocad import Autocad, APoint
import os

# 엑셀 파일 불러오기

excel_file = r"d:\work\_test_data.xlsx"
sheet = "AutoCAD_Export"
excel = comtypes.client.CreateObject("Excel.Application")
excel.Visible = False
#wb = excel.Workbooks.Open(excel_file)
wb = openpyxl.load_workbook(excel_file)
ws = wb.Sheets(sheet)


#excel = comtypes.client.CreateObject("Excel.Application")
#excel.Visible = False

#excel_file = r"d:\work\_test_data.xlsx"
#sheet = "AutoCAD_Export"

#wb = excel.Workbooks.Open(excel_path)
#excel.Visible = False


##wb = openpyxl.load_workbook("example.xlsx")
#wb = openpyxl.load_workbook(excel_file)
##ws = wb.active  # 활성 시트 가져오기
#ws = wb.Sheets(sheet)

# 특정 셀의 정렬 정보 가져오기 (예: A1 셀)
cell = ws["B3"]
alignment = cell.alignment

# 정렬 속성 출력
print(f"수평 정렬: {alignment.horizontal}")
print(f"수직 정렬: {alignment.vertical}")
print(f"텍스트 자동 줄 바꿈: {alignment.wrap_text}")
#print(f"셀 병합 여부: {alignment.mergeCells}")
print(f"셀 회전 각도: {alignment.text_rotation}")

# 파일 닫기
wb.close()