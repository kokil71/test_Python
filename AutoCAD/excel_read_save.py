import openpyxl
import os

def load_or_create_workbook(file_path, bln_data_only=True):
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path, data_only=bln_data_only)
        print(f"📂 기존 파일 열기: {file_path}")
    else:
        wb = openpyxl.Workbook()
        print(f"🆕 새 파일 생성: {file_path}")
    return wb

def get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"📄 기존 시트 열기: {sheet_name}")
    else:
        ws = wb.create_sheet(title=sheet_name)
        print(f"➕ 새 시트 생성: {sheet_name}")
    return ws

def read_sheet_data(ws):
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return data

def write_sheet_data(ws, data):
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def save_workbook(wb, file_path):
    wb.save(file_path)
    print(f"✅ 저장 완료: {file_path}")

# 사용 예제
file_path = r"d:\work\test_data.xlsx"
#file_path = "d:/work/test_data.xlsx"
sheet_name = "AutoCAD_Export"

# 1. 엑셀 열기 또는 생성
wb = load_or_create_workbook(file_path)

# 2. 시트 열기 또는 생성
ws = get_or_create_sheet(wb, sheet_name)

# 3. 기존 데이터 읽기
existing_data = read_sheet_data(ws)
print("📋 기존 데이터:")
for row in existing_data:
    print(row)

# 4. 새 데이터 쓰기
new_data = [
    ["항목", "수량", "단가", "금액"],
    ["철근", 10, 5000, 50000],
    ["콘크리트", 3, 120000, 360000],
]
write_sheet_data(ws, new_data)

# 5. 저장
save_workbook(wb, file_path)


# 6. 셀 주소, 행 번호, 열 번호 가져오기

# 엑셀 파일 열기
#wb = load_or_create_workbook(file_path)
#ws = get_or_create_sheet(wb, sheet_name)   # 쉬트

#wb = openpyxl.load_workbook("파일경로.xlsx")
#sheet = wb.active  # 또는 wb["시트이름"]

# 예: A1 셀 읽기
#cell = sheet["A1"]
cell = ws["A1"]

print("셀 주소:", cell.coordinate)  # 👉 A1
print("행 번호:", cell.row)         # 👉 1
print("열 번호:", cell.column)     # 👉 1 (열 A는 숫자 1)


alignment = cell.alignment

# 정렬 속성 출력
print(f"수평 정렬: {alignment.horizontal}")
print(f"수직 정렬: {alignment.vertical}")
print(f"텍스트 자동 줄 바꿈: {alignment.wrap_text}")
#print(f"셀 병합 여부: {alignment.mergeCells}")
print(f"셀 회전 각도: {alignment.text_rotation}")

# 2. 반복하면서 각 셀의 위치 가져오기
for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
    for cell in row:
        print(f"값: {cell.value}, 주소: {cell.coordinate}, 행: {cell.row}, 열: {cell.column}")

# 3. 열 번호 → 열 문자 변환 (openpyxl.utils.get_column_letter)
from openpyxl.utils import get_column_letter
print(get_column_letter(1))  # 👉 "A"
print(get_column_letter(28)) # 👉 "AB"

# 4. 열 문자 → 열 번호 변환 (openpyxl.utils.column_index_from_string)
from openpyxl.utils import column_index_from_string
print(column_index_from_string("A"))  # 👉 1
print(column_index_from_string("AB")) # 👉 28

# 5. 파일 닫기
# 암시적 정리
del wb
# win32com.client를 사용할 때 (명시적으로 닫기)
#wb.Close(SaveChanges=0)  # 0: 저장 안함, -1: 저장함
##excel.Quit()

# 6. 수식 가져오는 방법 (예: =SUM(A1:A3) 같은 수식)
# 수식을 가져오려면 반드시 data_only=False
file_path = r"d:\work\_test_data.xlsx"
bln_data_only = False                                                  # 수식으로 가져오기
#bln_data_only = True                                                    # 데이타로 가져오기
wb = load_or_create_workbook(file_path, bln_data_only)
ws = get_or_create_sheet(wb, sheet_name)
#wb = openpyxl.load_workbook("파일경로.xlsx", data_only=False)
#sheet = wb.active
cell = ws["B4"]  # 수식이 들어 있는 셀
print(f"B4 셀의 수식: {cell}")
print("B4 셀의 수식:", cell.value)  # 👉 예: "=SUM(A1:A3)"
#formula = ws["B4"].formula
#print(f"B4 셀의 수식: {formula}")
# 7. 파일 닫기
# 암시적 정리
#del wb

#8. 특정 셀의 정렬 정보 가져오기 (예: A1 셀)
#cell = ws["B3"]
alignment = cell.alignment
# 정렬 속성 출력
print(f"수평 정렬: {alignment.horizontal}")
print(f"수직 정렬: {alignment.vertical}")
print(f"텍스트 자동 줄 바꿈: {alignment.wrap_text}")
#print(f"셀 병합 여부: {alignment.mergeCells}")
print(f"셀 회전 각도: {alignment.text_rotation}")

# 파일 닫기
del wb