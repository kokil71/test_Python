# https://www.supplychaindataanalytics.com/working-with-autocad-in-excel-and-python/
# Working with AutoCAD in Excel and Python

import pythoncom
import win32gui
import win32com.client
import time
from pyautocad import Autocad, APoint, aDouble
import openpyxl as oex
#from math import pi

def connect_autocad():
    acad_app = None
    try:
        # 실행 중인 AutoCAD 인스턴스에 연결
        acad_app = win32com.client.GetActiveObject("AutoCAD.Application")        
        #print("✅ AutoCAD 실행 중입니다.")
        return Autocad(create_if_not_exists=False)          # pyautocad 연결
    except Exception:
        #print("❌ 실행 중인 AutoCAD 없음. 새 인스턴스를 시작합니다.")
        try:
            acad_app = win32com.client.Dispatch("AutoCAD.Application")
            #acad_app.Visible = True
            time.sleep(3)  # 로딩 대기
            #print("✅ 새 AutoCAD 인스턴스 시작 완료.")
            return Autocad(create_if_not_exists=False)      # pyautocad 연결
        except Exception as e:
            print("🚫 AutoCAD 연결 실패:", e)
            return None
    finally:
        if acad_app is not None:
            hwnd = win32gui.FindWindow(None, acad_app.Caption)
            bring_autocad_to_front(hwnd)      

def bring_autocad_to_front(hwnd):
    try:
        #acad = win32com.client.GetActiveObject("AutoCAD.Application")
        #hwnd = win32gui.FindWindow(None, acad.Caption)
        if hwnd:
            win32gui.ShowWindow(hwnd, 5)  # SW_SHOW
            win32gui.SetForegroundWindow(hwnd)
            #print("✅ AutoCAD 창을 전면으로 전환 완료.")
        else:
            print("❌ AutoCAD 창을 찾을 수 없습니다.")
    except Exception as e:
        print("🚫 오류:", e)

def draw_and_zoom(acad):    
    if acad is None:
        print("AutoCAD 연결이 안 되어 종료합니다.")
        return

    # Fetching data from Excel files for AutoCAD in Python
    # To fetch data from Excel files I need two variables.
    # The first one is used to store the workbook. 
    # The second variable is used to store the currently active worksheet. 
    # See below code for clarification.
    wbook = oex.load_workbook("D:\\work\\test_Python\\AutoCAD\\Python for AutoCAD\\test.xlsx")
    sh = wbook.active
    l1 = []
    for i in range(2, sh.max_row+1):
        for j in range(2, sh.max_column+1):
            cell = sh.cell(row=i, column=j)
            l1.append(cell.value)
    print(l1)

    # Creating AutoCAD polyline with pyautocad
    # Finally, I need to convert the list into a aDouble (Array of doubles) object. 
    # This allows me to use it as a input parameter when calling the AddPolyline pyautocad method.
    # This is the method that will actually draw the AutoCAD polyline. 
    # The code is displayed below.
    polygon = aDouble(l1)
    acad.model.AddPolyline(polygon)

    # 화면 갱신
    acad.app.Update()
    acad.app.ZoomExtents()


if __name__ == "__main__":    
    pythoncom.CoInitialize()  # COM 초기화
    acad = connect_autocad()    
    draw_and_zoom(acad)
    pythoncom.CoUninitialize()
